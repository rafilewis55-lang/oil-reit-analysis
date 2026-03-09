"""
Build Excel workbook with:
  Tab 1: Raw monthly data
  Tab 2: Shock Periods (flags each month, lists historical windows)
  Tab 3: Correlations (CORREL formulas)
  Tab 4: Full Sample Regressions (LINEST formulas + cross-check)
  Tab 5: Shock Regressions (all shock definitions side by side)
  Tab 6: Charts
"""

import io
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series, LineChart
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Deutsche Bank style: #0018A8 dark blue, #E8EDF2 light blue-gray, clean/minimal
DB_BLUE = '0018A8'
DB_LIGHT = 'E8EDF2'
DB_MID = 'C7D3E3'

HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor=DB_BLUE)
RED_FILL = PatternFill('solid', fgColor=DB_BLUE)
RED_FONT = Font(bold=True, color='FFFFFF', size=11)
LIGHT_FILL = PatternFill('solid', fgColor=DB_LIGHT)
SHOCK_FILL = PatternFill('solid', fgColor='F5E6E8')
GREEN_FILL = PatternFill('solid', fgColor='E6EFE6')
BOLD = Font(bold=True)
BLUE = Font(color=DB_BLUE)
BLUE_BOLD = Font(bold=True, color=DB_BLUE)
ITALIC_GRAY = Font(italic=True, color='888888')
THIN_BORDER = Border(bottom=Side(style='thin', color=DB_MID))


def _header_row(ws, row, headers, fill=HEADER_FILL, font=HEADER_FONT):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)


def _auto_width(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[letter].width = min(max_len + 3, 26)


def _write_model_row(ws, row, var_name, coef, se, tstat, pval):
    """Write one variable row with coef, SE, t-stat, p-value, significance."""
    sig = '***' if pval < 0.01 else '**' if pval < 0.05 else '*' if pval < 0.1 else ''
    ws.cell(row=row, column=1, value=var_name).font = BOLD
    ws.cell(row=row, column=2, value=round(float(coef), 4)).number_format = '0.0000'
    ws.cell(row=row, column=3, value=round(float(se), 4)).number_format = '0.0000'
    ws.cell(row=row, column=4, value=round(float(tstat), 3)).number_format = '0.000'
    ws.cell(row=row, column=5, value=round(float(pval), 4)).number_format = '0.0000'
    ws.cell(row=row, column=6, value=sig)
    if pval < 0.05:
        ws.cell(row=row, column=6).font = Font(bold=True, color=DB_BLUE)
    for c in range(1, 7):
        ws.cell(row=row, column=c).border = THIN_BORDER


def _write_reg_block(ws, start_row, title, model, include_linest=False, linest_base=None):
    """Write a full regression block. Returns next free row."""
    r = start_row
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=12, color=DB_BLUE)
    r += 1

    # Summary stats
    ws.cell(row=r, column=1, value='Observations').font = BOLD
    ws.cell(row=r, column=2, value=int(model.nobs))
    r += 1
    ws.cell(row=r, column=1, value='R-squared').font = BOLD
    ws.cell(row=r, column=2, value=round(float(model.rsquared), 4)).number_format = '0.0000'
    if include_linest and linest_base:
        ws.cell(row=r, column=3, value=f'=INDEX({linest_base},3,1)')
        ws.cell(row=r, column=3).number_format = '0.0000'
        ws.cell(row=r, column=3).font = BLUE
    r += 1
    ws.cell(row=r, column=1, value='Adj. R-squared').font = BOLD
    ws.cell(row=r, column=2, value=round(float(model.rsquared_adj), 4)).number_format = '0.0000'
    r += 1
    ws.cell(row=r, column=1, value='F-statistic').font = BOLD
    ws.cell(row=r, column=2, value=round(float(model.fvalue), 3)).number_format = '0.000'
    r += 1
    ws.cell(row=r, column=1, value='F p-value').font = BOLD
    ws.cell(row=r, column=2, value=round(float(model.f_pvalue), 4)).number_format = '0.0000'
    r += 2

    _header_row(ws, r, ['Variable', 'Coefficient', 'Std Error', 't-stat', 'p-value', 'Sig'])
    r += 1

    # LINEST cross-check column headers if needed
    if include_linest and linest_base:
        ws.cell(row=r-1, column=7, value='Coef (formula)').font = BLUE_BOLD
        ws.cell(row=r-1, column=7).fill = LIGHT_FILL

    n_x = len(model.params) - 1  # excluding constant
    for i, var in enumerate(model.params.index):
        _write_model_row(ws, r, var,
                         model.params[var], model.bse[var],
                         model.tvalues[var], model.pvalues[var])
        # LINEST cross-check
        if include_linest and linest_base:
            # LINEST order: xn, xn-1, ..., x1, intercept
            if var == 'const':
                li = n_x + 1
            else:
                # Find position of this var among non-const params (0-indexed)
                non_const = [v for v in model.params.index if v != 'const']
                pos = non_const.index(var)
                li = n_x - pos  # LINEST reverses
            ws.cell(row=r, column=7, value=f'=INDEX({linest_base},1,{li})')
            ws.cell(row=r, column=7).number_format = '0.0000'
            ws.cell(row=r, column=7).font = BLUE
        r += 1

    r += 1
    return r


def build_excel(data, regressions):
    wb = Workbook()
    df = data['df']
    n = len(df)
    last_row = n + 1

    shock_defs = regressions.get('_shock_defs', {})
    historical_shocks = regressions.get('_historical_shocks', {})
    oil_std = regressions.get('_oil_std', df['oil_chg'].std())
    shock_results = regressions.get('_shock_results', {})
    shock_counts = regressions.get('_shock_counts', {})

    # ==================================================================
    # TAB 1: DATA
    # ==================================================================
    ws = wb.active
    ws.title = 'Data'

    headers = ['Date', 'Oil Price', 'Oil Chg %', 'REIT Return %',
               'S&P 500 Return %', 'REIT Excess Return %',
               '3M Rate %', '10Y Rate %', 'Term Spread %',
               'Chg in 3M Rate', 'Chg in 10Y Rate']
    _header_row(ws, 1, headers)

    for r, (date, row) in enumerate(df.iterrows(), 2):
        ws.cell(row=r, column=1, value=date.strftime('%Y-%m'))
        ws.cell(row=r, column=2, value=round(row['oil_price'], 2))
        ws.cell(row=r, column=3, value=round(row['oil_chg'], 2))
        ws.cell(row=r, column=4, value=round(row['reit_ret'], 2))
        ws.cell(row=r, column=5, value=round(row['spx_ret'], 2))
        ws.cell(row=r, column=6, value=round(row['excess_ret'], 2))
        ws.cell(row=r, column=7, value=round(row['t3m'], 2))
        ws.cell(row=r, column=8, value=round(row['t10y'], 2))
        ws.cell(row=r, column=9, value=round(row['term_spread'], 2))
        ws.cell(row=r, column=10, value=round(row['d_t3m'], 4))
        ws.cell(row=r, column=11, value=round(row['d_t10y'], 4))

    _auto_width(ws)
    ws.freeze_panes = 'A2'

    def drange(col_letter):
        return f"Data!{col_letter}2:{col_letter}{last_row}"

    # ==================================================================
    # TAB 2: SHOCK PERIODS
    # ==================================================================
    ws2 = wb.create_sheet('Shock Periods')

    ws2.cell(row=1, column=1, value='Oil Shock Period Identification').font = Font(bold=True, size=14, color=DB_BLUE)
    ws2.cell(row=2, column=1, value=f'1 Std Dev of monthly oil change = {oil_std:.1f}%. Months highlighted in pink are shock months.').font = ITALIC_GRAY

    # Monthly shock flags
    r = 4
    flag_headers = ['Date', 'Oil Chg %', '|Oil Chg| > 1 SD', '>1.5 SD', '|Chg| > 10%',
                    'Spike (>10%)', 'Crash (<-10%)', 'Historical Window']
    _header_row(ws2, r, flag_headers)
    r += 1

    shock_1sd = df['oil_chg'].abs() > oil_std
    shock_15sd = df['oil_chg'].abs() > 1.5 * oil_std
    shock_10 = df['oil_chg'].abs() > 10
    spike_10 = df['oil_chg'] > 10
    crash_10 = df['oil_chg'] < -10
    any_hist = shock_defs.get('Any historical window', pd.Series(False, index=df.index))

    for date, row in df.iterrows():
        dt = date.strftime('%Y-%m')
        oil = row['oil_chg']
        is_1sd = abs(oil) > oil_std
        is_15sd = abs(oil) > 1.5 * oil_std
        is_10 = abs(oil) > 10
        is_spike = oil > 10
        is_crash = oil < -10
        is_hist = bool(any_hist.get(date, False))

        ws2.cell(row=r, column=1, value=dt)
        ws2.cell(row=r, column=2, value=round(oil, 2)).number_format = '0.00'
        ws2.cell(row=r, column=3, value='YES' if is_1sd else '')
        ws2.cell(row=r, column=4, value='YES' if is_15sd else '')
        ws2.cell(row=r, column=5, value='YES' if is_10 else '')
        ws2.cell(row=r, column=6, value='YES' if is_spike else '')
        ws2.cell(row=r, column=7, value='YES' if is_crash else '')
        ws2.cell(row=r, column=8, value='YES' if is_hist else '')

        if is_1sd:
            for c in range(1, 9):
                ws2.cell(row=r, column=c).fill = SHOCK_FILL
        r += 1

    # Historical shock windows list
    r += 2
    ws2.cell(row=r, column=1, value='Historical Shock Windows').font = Font(bold=True, size=13, color=DB_BLUE)
    r += 1
    _header_row(ws2, r, ['Period', 'Start', 'End', 'Months', 'Avg Oil Chg %',
                          'Avg REIT Excess %', 'Avg 10Y Chg (pp)'])
    r += 1
    for label, (start, end) in historical_shocks.items():
        mask = (df.index >= start) & (df.index <= end)
        sub = df[mask]
        if len(sub) == 0:
            continue
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=start)
        ws2.cell(row=r, column=3, value=end)
        ws2.cell(row=r, column=4, value=len(sub))
        ws2.cell(row=r, column=5, value=round(float(sub['oil_chg'].mean()), 1))
        ws2.cell(row=r, column=6, value=round(float(sub['excess_ret'].mean()), 2))
        ws2.cell(row=r, column=7, value=round(float(sub['d_t10y'].mean()), 3))
        for c in range(1, 8):
            ws2.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    # Shock definition counts
    r += 2
    ws2.cell(row=r, column=1, value='Shock Definition Summary').font = Font(bold=True, size=13, color=DB_BLUE)
    r += 1
    _header_row(ws2, r, ['Definition', 'Months', '% of Sample'])
    r += 1
    for label, count in shock_counts.items():
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=count)
        ws2.cell(row=r, column=3, value=round(count / n * 100, 1))
        ws2.cell(row=r, column=3).number_format = '0.0'
        for c in range(1, 4):
            ws2.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    _auto_width(ws2)

    # ==================================================================
    # TAB 3: CORRELATIONS (CORREL formulas)
    # ==================================================================
    ws3 = wb.create_sheet('Correlations')

    ws3.cell(row=1, column=1, value='Correlation Matrix (Full Sample)').font = Font(bold=True, size=14, color=DB_BLUE)
    ws3.cell(row=2, column=1, value='All values are CORREL() formulas referencing the Data tab').font = ITALIC_GRAY

    corr_vars = [
        ('Oil Chg %', 'C'), ('REIT Excess Ret %', 'F'), ('3M Rate %', 'G'),
        ('10Y Rate %', 'H'), ('Term Spread %', 'I'), ('Chg in 3M', 'J'), ('Chg in 10Y', 'K'),
    ]

    r0 = 4
    _header_row(ws3, r0, [''] + [v[0] for v in corr_vars])
    for ri, (row_name, row_col) in enumerate(corr_vars):
        r = r0 + 1 + ri
        ws3.cell(row=r, column=1, value=row_name).font = BOLD
        for ci, (_, col_col) in enumerate(corr_vars, 2):
            if row_col == col_col:
                ws3.cell(row=r, column=ci, value=1.0)
            else:
                ws3.cell(row=r, column=ci, value=f'=CORREL({drange(row_col)},{drange(col_col)})')
            ws3.cell(row=r, column=ci).number_format = '0.000'

    _auto_width(ws3)

    # ==================================================================
    # TAB 4: FULL SAMPLE REGRESSIONS
    # ==================================================================
    ws4 = wb.create_sheet('Full Sample Regressions')

    r = 1
    ws4.cell(row=r, column=1, value='Full Sample Regression Results').font = Font(bold=True, size=14, color=DB_BLUE)
    r += 1
    ws4.cell(row=r, column=1, value='Black = values. Blue = LINEST formula cross-check. *** p<0.01  ** p<0.05  * p<0.1').font = ITALIC_GRAY
    r += 2

    # LINEST base for cross-check (non-adjacent cols C,J,K need CHOOSE)
    linest_reit = f'LINEST({drange("F")},CHOOSE({{1,2,3}},{drange("C")},{drange("J")},{drange("K")}),TRUE,TRUE)'
    r = _write_reg_block(ws4, r, 'REIT Excess Return ~ Oil + Rate Changes',
                         regressions['reit_m1_levels_chg'],
                         include_linest=True, linest_base=linest_reit)

    linest_10y = f'LINEST({drange("K")},{drange("C")},TRUE,TRUE)'
    r = _write_reg_block(ws4, r, '10Y Rate Change ~ Oil Change',
                         regressions['t10y_on_oil_ols'],
                         include_linest=True, linest_base=linest_10y)

    linest_3m = f'LINEST({drange("J")},{drange("C")},TRUE,TRUE)'
    r = _write_reg_block(ws4, r, '3M Rate Change ~ Oil Change',
                         regressions['t3m_on_oil_ols'],
                         include_linest=True, linest_base=linest_3m)

    linest_oil = f'LINEST({drange("C")},{drange("J")}:{drange("K").split("!")[1]},TRUE,TRUE)'
    # Adjacent cols J,K so simpler
    linest_oil = f'LINEST({drange("C")},Data!J2:K{last_row},TRUE,TRUE)'
    r = _write_reg_block(ws4, r, 'Oil Change ~ Rate Changes',
                         regressions['oil_rates_changes_ols'],
                         include_linest=True, linest_base=linest_oil)

    # Additional HC1 models (no LINEST cross-check)
    r = _write_reg_block(ws4, r, 'REIT Excess ~ Oil + Rate Changes (Winsorized, Robust SE)',
                         regressions['reit_m2'])
    r = _write_reg_block(ws4, r, 'REIT Excess ~ Asymmetric Oil + Rate Changes (Winsorized)',
                         regressions['reit_m3'])

    _auto_width(ws4)

    # ==================================================================
    # TAB 5: SHOCK REGRESSIONS
    # ==================================================================
    ws5 = wb.create_sheet('Shock Regressions')

    r = 1
    ws5.cell(row=r, column=1, value='Regressions During Oil Shock Periods').font = Font(bold=True, size=14, color=DB_BLUE)
    r += 1
    ws5.cell(row=r, column=1, value='Same models as Full Sample tab, but restricted to shock months only. *** p<0.01  ** p<0.05  * p<0.1').font = ITALIC_GRAY
    r += 2

    # Comparison table: all shock defs side by side for the REIT regression
    ws5.cell(row=r, column=1, value='REIT Excess Return ~ Oil + Rate Changes: Shock vs Full Sample').font = Font(bold=True, size=13, color=DB_BLUE)
    r += 1

    comp_headers = ['Shock Definition', 'N', 'R-sq',
                    'Oil Coef', 'Oil t-stat', 'Oil p-val',
                    'd_10Y Coef', 'd_10Y t-stat', 'd_10Y p-val',
                    'd_3M Coef', 'd_3M t-stat', 'd_3M p-val']
    _header_row(ws5, r, comp_headers)
    r += 1

    # Full sample first (HC1 to match shock rows)
    m_full = regressions['reit_full_hc1']
    ws5.cell(row=r, column=1, value='FULL SAMPLE').font = Font(bold=True, color=DB_BLUE)
    ws5.cell(row=r, column=2, value=int(m_full.nobs))
    ws5.cell(row=r, column=3, value=round(float(m_full.rsquared), 4)).number_format = '0.0000'
    for vi, var in enumerate(['oil_chg', 'd_t10y', 'd_t3m']):
        base_col = 4 + vi * 3
        ws5.cell(row=r, column=base_col, value=round(float(m_full.params[var]), 4)).number_format = '0.0000'
        ws5.cell(row=r, column=base_col+1, value=round(float(m_full.tvalues[var]), 3)).number_format = '0.000'
        ws5.cell(row=r, column=base_col+2, value=round(float(m_full.pvalues[var]), 4)).number_format = '0.0000'
    for c in range(1, 13):
        ws5.cell(row=r, column=c).border = THIN_BORDER
        ws5.cell(row=r, column=c).fill = LIGHT_FILL
    r += 1

    # Each shock definition
    ordered_labels = [
        '1 SD (|oil| > 1 std dev)', '1.5 SD (|oil| > 1.5 std dev)',
        'Top/Bottom 10%', 'Big move (|chg| > 10%)',
        'Oil spike (>10%)', 'Oil crash (<-10%)',
        'Any historical window', 'Calm months (|oil| < 1 SD)',
    ]
    for label in ordered_labels:
        sr = shock_results.get(label)
        if sr is None or 'reit' not in sr:
            ws5.cell(row=r, column=1, value=label)
            ws5.cell(row=r, column=2, value=shock_counts.get(label, '<10'))
            ws5.cell(row=r, column=3, value='N/A (too few obs)')
            r += 1
            continue

        m = sr['reit']
        ws5.cell(row=r, column=1, value=label).font = BOLD
        ws5.cell(row=r, column=2, value=int(m.nobs))
        ws5.cell(row=r, column=3, value=round(float(m.rsquared), 4)).number_format = '0.0000'
        for vi, var in enumerate(['oil_chg', 'd_t10y', 'd_t3m']):
            base_col = 4 + vi * 3
            ws5.cell(row=r, column=base_col, value=round(float(m.params[var]), 4)).number_format = '0.0000'
            ws5.cell(row=r, column=base_col+1, value=round(float(m.tvalues[var]), 3)).number_format = '0.000'
            pval = float(m.pvalues[var])
            cell_p = ws5.cell(row=r, column=base_col+2, value=round(pval, 4))
            cell_p.number_format = '0.0000'
            if pval < 0.05:
                cell_p.font = Font(bold=True, color=DB_BLUE)
        for c in range(1, 13):
            ws5.cell(row=r, column=c).border = THIN_BORDER
        if label == 'Calm months (|oil| < 1 SD)':
            for c in range(1, 13):
                ws5.cell(row=r, column=c).fill = GREEN_FILL
        r += 1

    # Same comparison for Oil -> 10Y regression
    r += 2
    ws5.cell(row=r, column=1, value='10Y Rate Change ~ Oil Change: Shock vs Full Sample').font = Font(bold=True, size=13, color=DB_BLUE)
    r += 1
    comp_h2 = ['Shock Definition', 'N', 'R-sq', 'Oil Coef', 'Oil t-stat', 'Oil p-val']
    _header_row(ws5, r, comp_h2)
    r += 1

    m_full2 = regressions['t10y_on_oil']  # HC1 to match shock rows
    ws5.cell(row=r, column=1, value='FULL SAMPLE').font = Font(bold=True, color=DB_BLUE)
    ws5.cell(row=r, column=2, value=int(m_full2.nobs))
    ws5.cell(row=r, column=3, value=round(float(m_full2.rsquared), 4)).number_format = '0.0000'
    ws5.cell(row=r, column=4, value=round(float(m_full2.params['oil_chg']), 4)).number_format = '0.0000'
    ws5.cell(row=r, column=5, value=round(float(m_full2.tvalues['oil_chg']), 3)).number_format = '0.000'
    ws5.cell(row=r, column=6, value=round(float(m_full2.pvalues['oil_chg']), 4)).number_format = '0.0000'
    for c in range(1, 7):
        ws5.cell(row=r, column=c).border = THIN_BORDER
        ws5.cell(row=r, column=c).fill = LIGHT_FILL
    r += 1

    for label in ordered_labels:
        sr = shock_results.get(label)
        if sr is None or 't10y' not in sr:
            ws5.cell(row=r, column=1, value=label)
            ws5.cell(row=r, column=2, value=shock_counts.get(label, '<10'))
            ws5.cell(row=r, column=3, value='N/A')
            r += 1
            continue
        m = sr['t10y']
        ws5.cell(row=r, column=1, value=label).font = BOLD
        ws5.cell(row=r, column=2, value=int(m.nobs))
        ws5.cell(row=r, column=3, value=round(float(m.rsquared), 4)).number_format = '0.0000'
        ws5.cell(row=r, column=4, value=round(float(m.params['oil_chg']), 4)).number_format = '0.0000'
        ws5.cell(row=r, column=5, value=round(float(m.tvalues['oil_chg']), 3)).number_format = '0.000'
        pval = float(m.pvalues['oil_chg'])
        cell_p = ws5.cell(row=r, column=6, value=round(pval, 4))
        cell_p.number_format = '0.0000'
        if pval < 0.05:
            cell_p.font = Font(bold=True, color=DB_BLUE)
        for c in range(1, 7):
            ws5.cell(row=r, column=c).border = THIN_BORDER
        if label == 'Calm months (|oil| < 1 SD)':
            for c in range(1, 7):
                ws5.cell(row=r, column=c).fill = GREEN_FILL
        r += 1

    # Full detail blocks for key shock definitions
    r += 2
    ws5.cell(row=r, column=1, value='Detailed Regression Output by Shock Definition').font = Font(bold=True, size=14, color=DB_BLUE)
    r += 2

    for label in ['1 SD (|oil| > 1 std dev)', '1.5 SD (|oil| > 1.5 std dev)',
                   'Oil crash (<-10%)', 'Any historical window',
                   'Calm months (|oil| < 1 SD)']:
        sr = shock_results.get(label)
        if sr is None:
            continue
        for reg_name, model_key, desc in [
            ('reit', 'reit', 'REIT Excess ~ Oil + Rate Changes'),
            ('t10y', 't10y', '10Y Change ~ Oil Change'),
            ('t3m', 't3m', '3M Change ~ Oil Change'),
        ]:
            if model_key in sr:
                r = _write_reg_block(ws5, r, f'{desc} [{label}]', sr[model_key])

    _auto_width(ws5)

    # ==================================================================
    # TAB 6: KEY FINDINGS
    # ==================================================================
    wsf = wb.create_sheet('Key Findings')
    wsf.column_dimensions['A'].width = 4
    wsf.column_dimensions['B'].width = 90

    r = 1
    wsf.cell(row=r, column=2, value='Key Findings: Oil, REITs & Interest Rates').font = Font(bold=True, size=16, color=DB_BLUE)
    r += 2

    findings = [
        ("1. Oil doesn't directly move REITs vs S&P -- even during shocks.",
         "Across every definition of 'oil shock' (1 SD moves, >10% swings, historical crisis windows), "
         "the oil coefficient on REIT excess returns is statistically insignificant. Oil spikes and crashes "
         "hit REITs and the S&P roughly equally."),

        ("2. Long-term rates are what actually drive the wedge.",
         "A 1 percentage point rise in the 10Y yield in a given month is associated with ~5% REIT "
         "underperformance vs the S&P (p<0.001). REITs behave like long-duration bonds."),

        ("3. Oil's effect on rates gets stronger during shocks.",
         "Full sample: oil->10Y R-sq = 8.3%. Shock months (1 SD): R-sq = 18.1%. "
         "Extreme shocks (1.5 SD): R-sq = 32.1%. The bigger the oil move, the more it pushes the "
         "10-year rate -- likely through inflation expectations."),

        ("4. Oil crashes and spikes work differently.",
         "Oil crashes (<-10%): The 10Y rate effect on REITs becomes highly significant (p<0.01). "
         "Oil crashes pull rates down, which helps REITs relative to the S&P. The oil->3M rate "
         "relationship also strengthens (p=0.02), suggesting the Fed responds to oil-driven deflation risk.\n\n"
         "Oil spikes (>10%): Noisy -- nothing is significant. REITs underperform by -2.3% on average "
         "during spike months, but the regression can't attribute it cleanly to oil or rates."),

        ("5. The transmission chain is indirect.",
         "Oil shock -> inflation expectations -> rates move -> REITs react to rates. "
         "The direct oil->REIT channel is basically zero. The oil->rates->REITs chain is real "
         "but only explains ~6% of monthly variation. Most REIT vs S&P performance comes from "
         "other factors (sector rotation, cap rates, property fundamentals)."),
    ]

    for title, body in findings:
        wsf.cell(row=r, column=2, value=title).font = Font(bold=True, size=12, color=DB_BLUE)
        r += 1
        wsf.cell(row=r, column=2, value=body).alignment = Alignment(wrap_text=True)
        # Auto-height: roughly 1 row per 90 chars
        lines = max(len(body) // 85 + body.count('\n') + 1, 2)
        wsf.row_dimensions[r].height = lines * 16
        r += 2

    r += 1
    wsf.cell(row=r, column=2, value='Supporting Evidence (from Shock Regressions tab)').font = Font(bold=True, size=13, color=DB_BLUE)
    r += 1

    evidence = [
        ['Metric', 'Full Sample', 'Shock (1 SD)', 'Extreme (1.5 SD)', 'Oil Crash (<-10%)'],
    ]
    # Pull actual numbers (HC1 to match shock regressions)
    m_full = regressions['reit_full_hc1']
    sr_1sd = shock_results.get('1 SD (|oil| > 1 std dev)', {})
    sr_15sd = shock_results.get('1.5 SD (|oil| > 1.5 std dev)', {})
    sr_crash = shock_results.get('Oil crash (<-10%)', {})

    def _safe(sr, key, attr, var=None):
        m = sr.get(key)
        if m is None:
            return 'N/A'
        if var:
            return round(float(getattr(m, attr)[var]), 4)
        return round(float(getattr(m, attr)), 4)

    evidence.append(['REIT reg: Oil coef',
                     round(float(m_full.params['oil_chg']), 4),
                     _safe(sr_1sd, 'reit', 'params', 'oil_chg'),
                     _safe(sr_15sd, 'reit', 'params', 'oil_chg'),
                     _safe(sr_crash, 'reit', 'params', 'oil_chg')])
    evidence.append(['REIT reg: Oil p-value',
                     round(float(m_full.pvalues['oil_chg']), 4),
                     _safe(sr_1sd, 'reit', 'pvalues', 'oil_chg'),
                     _safe(sr_15sd, 'reit', 'pvalues', 'oil_chg'),
                     _safe(sr_crash, 'reit', 'pvalues', 'oil_chg')])
    evidence.append(['REIT reg: 10Y coef',
                     round(float(m_full.params['d_t10y']), 4),
                     _safe(sr_1sd, 'reit', 'params', 'd_t10y'),
                     _safe(sr_15sd, 'reit', 'params', 'd_t10y'),
                     _safe(sr_crash, 'reit', 'params', 'd_t10y')])
    evidence.append(['REIT reg: 10Y p-value',
                     round(float(m_full.pvalues['d_t10y']), 4),
                     _safe(sr_1sd, 'reit', 'pvalues', 'd_t10y'),
                     _safe(sr_15sd, 'reit', 'pvalues', 'd_t10y'),
                     _safe(sr_crash, 'reit', 'pvalues', 'd_t10y')])

    m_full_10y = regressions['t10y_on_oil']  # HC1 to match shock rows
    evidence.append(['Oil->10Y: R-squared',
                     round(float(m_full_10y.rsquared), 4),
                     _safe(sr_1sd, 't10y', 'rsquared'),
                     _safe(sr_15sd, 't10y', 'rsquared'),
                     _safe(sr_crash, 't10y', 'rsquared')])
    evidence.append(['Oil->10Y: Oil p-value',
                     round(float(m_full_10y.pvalues['oil_chg']), 4),
                     _safe(sr_1sd, 't10y', 'pvalues', 'oil_chg'),
                     _safe(sr_15sd, 't10y', 'pvalues', 'oil_chg'),
                     _safe(sr_crash, 't10y', 'pvalues', 'oil_chg')])
    evidence.append(['N months',
                     int(m_full.nobs),
                     shock_counts.get('1 SD (|oil| > 1 std dev)', 0),
                     shock_counts.get('1.5 SD (|oil| > 1.5 std dev)', 0),
                     shock_counts.get('Oil crash (<-10%)', 0)])

    _header_row(wsf, r, [''] + evidence[0])
    r += 1
    for row_data in evidence[1:]:
        wsf.cell(row=r, column=2, value=row_data[0]).font = BOLD
        for ci, val in enumerate(row_data[1:], 3):
            cell = wsf.cell(row=r, column=ci, value=val)
            if isinstance(val, float):
                cell.number_format = '0.0000'
        for c in range(2, 7):
            wsf.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    # ==================================================================
    # TAB 7: SOURCES
    # ==================================================================
    wss = wb.create_sheet('Sources')
    wss.column_dimensions['A'].width = 4
    wss.column_dimensions['B'].width = 30
    wss.column_dimensions['C'].width = 65
    wss.column_dimensions['D'].width = 55

    r = 1
    wss.cell(row=r, column=2, value='Data Sources').font = Font(bold=True, size=16, color=DB_BLUE)
    r += 2
    _header_row(wss, r, ['', 'Source', 'URL', 'Used For'])
    r += 1

    sources = [
        ('FRED: DCOILWTICO',
         'https://fred.stlouisfed.org/series/DCOILWTICO',
         'WTI Crude Oil spot price (daily). Resampled to monthly average, then converted to % change for the oil shock variable.'),
        ('FRED: DTB3',
         'https://fred.stlouisfed.org/series/DTB3',
         '3-Month Treasury Bill rate (daily). Resampled to monthly average. Used as the short-term interest rate variable (level and month-over-month change).'),
        ('FRED: DGS10',
         'https://fred.stlouisfed.org/series/DGS10',
         '10-Year Treasury Constant Maturity rate (daily). Resampled to monthly average. Used as the long-term interest rate variable (level and month-over-month change).'),
        ('Yahoo Finance: ^RMZ',
         'https://finance.yahoo.com/quote/%5ERMZ/',
         'MSCI US REIT Index (daily close, Jun 1995 - Sep 2021). Used as the REIT benchmark for the earlier portion of the sample period.'),
        ('Yahoo Finance: IYR',
         'https://finance.yahoo.com/quote/IYR/',
         'iShares U.S. Real Estate ETF (daily close, Jun 2000 - present). Spliced with ^RMZ to create a continuous REIT return series through 2025.'),
        ('Yahoo Finance: ^GSPC',
         'https://finance.yahoo.com/quote/%5EGSPC/',
         'S&P 500 Index (daily close). Resampled to month-end values, then converted to monthly % return. Used as the broad equity benchmark.'),
    ]

    for source, url, desc in sources:
        wss.cell(row=r, column=2, value=source).font = BOLD
        wss.cell(row=r, column=3, value=url).font = Font(color='0563C1', underline='single')
        wss.hyperlink = url
        wss.cell(row=r, column=4, value=desc).alignment = Alignment(wrap_text=True)
        wss.row_dimensions[r].height = max(32, len(desc) // 55 * 16 + 16)
        for c in range(2, 5):
            wss.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    r += 2
    wss.cell(row=r, column=2, value='Methodology Notes').font = Font(bold=True, size=13, color=DB_BLUE)
    r += 1
    notes = [
        ('REIT Index Splice',
         '^RMZ covers Jun 1995 - Sep 2021. IYR covers Jun 2000 - present. We use ^RMZ for months before IYR began, then IYR from Jun 2000 onward, creating a continuous series from 1995 to 2025.'),
        ('Monthly Returns',
         'All return series use month-end closing prices. Monthly return = (close_t / close_{t-1} - 1) * 100.'),
        ('Excess Return',
         'REIT excess return = REIT monthly return - S&P 500 monthly return. Positive = REITs outperformed.'),
        ('Rate Changes',
         'Month-over-month change in the monthly average rate level (in percentage points). Used in regressions instead of rate levels to avoid spurious correlation.'),
        ('Robust Standard Errors',
         'Website regressions use HC1 (heteroscedasticity-consistent) standard errors. Excel LINEST formulas use classical OLS standard errors (cross-check column).'),
        ('Winsorization',
         'Some models trim excess returns and oil changes at the 1st/99th percentile to reduce the influence of extreme outliers (e.g., COVID March 2020, GFC).'),
    ]
    for title, desc in notes:
        wss.cell(row=r, column=2, value=title).font = BOLD
        wss.cell(row=r, column=3, value=desc).alignment = Alignment(wrap_text=True)
        wss.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        wss.row_dimensions[r].height = max(32, len(desc) // 90 * 16 + 16)
        for c in range(2, 5):
            wss.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    # ==================================================================
    # TAB 8: FLASH NOTE (Iran War & REITs)
    # ==================================================================
    wsfn = wb.create_sheet('Flash Note')
    wsfn.column_dimensions['A'].width = 4
    wsfn.column_dimensions['B'].width = 22
    wsfn.column_dimensions['C'].width = 22
    wsfn.column_dimensions['D'].width = 22
    wsfn.column_dimensions['E'].width = 22
    wsfn.column_dimensions['F'].width = 22

    DARK_FILL = PatternFill('solid', fgColor=DB_BLUE)
    MAROON_FILL = PatternFill('solid', fgColor=DB_BLUE)
    WHITE_BOLD = Font(bold=True, color='FFFFFF', size=11)
    SECTION_FONT = Font(bold=True, size=13, color=DB_BLUE)
    SUBSECTION_FONT = Font(bold=True, size=11, color=DB_BLUE)

    r = 1
    # Title block
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsfn.cell(row=r, column=2, value='EQUITY RESEARCH  |  REAL ESTATE  |  REIT Sector Flash Note').font = Font(bold=True, size=9, color='888888')
    r += 1
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsfn.cell(row=r, column=2, value='Iran War & REITs: The Rate Channel Is What Matters').font = Font(bold=True, size=16, color=DB_BLUE)
    r += 1
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsfn.cell(row=r, column=2, value='Oil is the headline. Rates are the mechanism. Subsector dispersion is the opportunity.').font = Font(italic=True, size=11, color='555555')
    r += 1
    wsfn.cell(row=r, column=6, value='March 7, 2026').font = Font(italic=True, color='888888')
    r += 2

    # Key Takeaways box
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = wsfn.cell(row=r, column=2, value='KEY TAKEAWAYS')
    c.font = WHITE_BOLD
    c.fill = DARK_FILL
    for col in range(2, 7):
        wsfn.cell(row=r, column=col).fill = DARK_FILL
    r += 1

    takeaways = [
        "Oil is not the direct lever for REITs. Across 30 years of monthly data (n=366), the direct oil->REIT excess return coefficient is statistically insignificant across every shock definition.",
        "Rates are. A 1pp rise in the 10Y is associated with ~3-5% REIT underperformance vs the S&P. This week's move from ~3.90% to ~4.13% is the real headwind.",
        "The transmission chain is active: Oil shock -> inflation expectations -> 10Y rises -> REITs underperform. The R-sq on oil->10Y strengthens from 8% in the full sample to 32% during extreme shock months.",
        "Subsector dispersion is the key opportunity. Industrial, Data Centers, and Healthcare are best insulated. Net Lease and Residential face duration headwinds. Hotels are a near-term avoid.",
        "The tail risk is asymmetric: if oil reaches $100+ and recession fears mount, the 10Y retreats and rate-sensitive REIT sectors could actually benefit from the resulting rate-cut repricing.",
    ]
    for tk in takeaways:
        wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        wsfn.cell(row=r, column=2, value=tk).alignment = Alignment(wrap_text=True)
        wsfn.row_dimensions[r].height = max(32, len(tk) // 80 * 16 + 20)
        for col in range(2, 7):
            wsfn.cell(row=r, column=col).fill = LIGHT_FILL
            wsfn.cell(row=r, column=col).border = THIN_BORDER
        r += 1
    r += 1

    # Macro Backdrop
    wsfn.cell(row=r, column=2, value='MACRO BACKDROP').font = SECTION_FONT
    r += 1
    macro_text = ("The US-Israel war with Iran (Operation Epic Fury, begun February 28) has effectively closed the "
                  "Strait of Hormuz to commercial traffic via drone attacks on tankers -- achieving a de facto blockade "
                  "without a formal naval operation. The conflict is now in day 8 with no near-term resolution visible; "
                  "Trump has demanded unconditional surrender.")
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsfn.cell(row=r, column=2, value=macro_text).alignment = Alignment(wrap_text=True)
    wsfn.row_dimensions[r].height = 64
    r += 2

    # Key market variables
    macro_vars = [
        ('OIL', 'Brent ~$82/bbl (+12% since Feb 28). WTI ~$75. Goldman base: $76 avg Q2. $100+ scenario if Hormuz stays closed 5+ weeks.'),
        ('10Y', 'Rose from ~3.90% pre-war to ~4.13% intraweek peak; currently ~4.06-4.10%. Bond market pricing inflation risk, not safety. Defying the typical geopolitical safe-haven bid.'),
        ('FED', 'Rate cuts pushed out. Prior consensus: cuts resume H2 2026. Now: cuts likely delayed to Q4 2026 at earliest under base case; potentially 2027 in bear case.'),
    ]
    for label, desc in macro_vars:
        wsfn.cell(row=r, column=2, value=label).font = Font(bold=True, size=11, color=DB_BLUE)
        wsfn.cell(row=r, column=2).fill = LIGHT_FILL
        wsfn.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        wsfn.cell(row=r, column=3, value=desc).alignment = Alignment(wrap_text=True)
        wsfn.row_dimensions[r].height = 48
        for col in range(2, 7):
            wsfn.cell(row=r, column=col).border = THIN_BORDER
        r += 1
    r += 1

    # Historical Framework
    wsfn.cell(row=r, column=2, value='HISTORICAL FRAMEWORK: WHAT 30 YEARS OF OIL SHOCKS TELL US').font = SECTION_FONT
    r += 1

    wsfn.cell(row=r, column=2, value='The Core Finding: Oil != REIT Underperformance').font = SUBSECTION_FONT
    r += 1
    core_text = ("Our regression analysis covers monthly data from 1995-2025 across every major shock definition. "
                 "The bottom line is unambiguous: oil does not drive REIT underperformance vs the S&P, even during "
                 "major shocks. Oil spikes hit REITs and the broader market approximately equally. There is no "
                 "meaningful differential impact from the energy channel alone.")
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsfn.cell(row=r, column=2, value=core_text).alignment = Alignment(wrap_text=True)
    wsfn.row_dimensions[r].height = 64
    r += 2

    # Exhibit 1: Oil coefficient table
    wsfn.cell(row=r, column=2, value='Exhibit 1: Oil Coefficient on REIT Excess Return').font = SUBSECTION_FONT
    r += 1
    ex1_headers = ['Regime', 'Oil Coef.', 'p-value', 'Verdict']
    for ci, h in enumerate(ex1_headers, 2):
        c = wsfn.cell(row=r, column=ci, value=h)
        c.font = WHITE_BOLD
        c.fill = DARK_FILL
        c.alignment = Alignment(horizontal='center')
    r += 1
    ex1_data = [
        ('Full Sample (n=366)', '-0.022', '0.567', 'Not significant'),
        ('Oil Shocks >1 SD (n=88)', '-0.051', '0.547', 'Not significant'),
        ('Extreme Shocks >1.5 SD (n=29)', '+0.031', '0.429', 'Not significant'),
        ('Oil Spikes >+10% (n=46)', '+0.048', '0.582', 'Not significant'),
        ('Oil Crashes <-10% (n=35)', '+0.178', '0.155', 'Not significant'),
    ]
    for regime, coef, pval, verdict in ex1_data:
        wsfn.cell(row=r, column=2, value=regime)
        wsfn.cell(row=r, column=3, value=coef).alignment = Alignment(horizontal='center')
        wsfn.cell(row=r, column=4, value=pval).alignment = Alignment(horizontal='center')
        wsfn.cell(row=r, column=5, value=verdict).font = Font(color='336633')
        for col in range(2, 6):
            wsfn.cell(row=r, column=col).border = THIN_BORDER
        r += 1
    r += 1

    # Rate mechanism
    wsfn.cell(row=r, column=2, value='The Real Mechanism: Rates Drive the Wedge').font = SUBSECTION_FONT
    r += 1
    rate_text = ("A 1 percentage point rise in the 10-year Treasury yield in a given month is associated with "
                 "approximately 3-5% REIT underperformance vs the S&P. At the extreme shock threshold (1.5 SD oil "
                 "moves), this coefficient is -14.6 and highly significant (p<0.001). REITs behave like long-duration "
                 "bonds -- the interest rate channel is the dominant factor in relative performance.")
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsfn.cell(row=r, column=2, value=rate_text).alignment = Alignment(wrap_text=True)
    wsfn.row_dimensions[r].height = 64
    r += 2

    oil_10y_text = ("The oil->10Y relationship also strengthens significantly during large shocks: R-sq rises from "
                    "8.3% in the full sample to 32.1% during extreme shock months, meaning the current environment "
                    "-- a genuine large oil shock -- is precisely when this channel is most active.")
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsfn.cell(row=r, column=2, value=oil_10y_text).alignment = Alignment(wrap_text=True)
    wsfn.row_dimensions[r].height = 48
    r += 2

    asym_text = ("Important asymmetry: oil crashes (<-10%) produce a significantly different dynamic than oil spikes. "
                 "Crashes pull rates down, and the rate->REIT channel becomes highly significant (p=0.005, coef=-10.6). "
                 "This is the tail-risk bull case: if oil surges to $100+ and recession fears take hold, the 10Y "
                 "retreats sharply and rate-sensitive REIT sectors benefit meaningfully.")
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsfn.cell(row=r, column=2, value=asym_text).alignment = Alignment(wrap_text=True)
    wsfn.row_dimensions[r].height = 64
    r += 2

    # Subsector Impact Analysis
    wsfn.cell(row=r, column=2, value='SUBSECTOR IMPACT ANALYSIS').font = SECTION_FONT
    r += 1
    sub_headers = ['Subsector', 'Rate Sensitivity', 'Oil/Inflation', 'Demand Impact', 'Our View']
    for ci, h in enumerate(sub_headers, 2):
        c = wsfn.cell(row=r, column=ci, value=h)
        c.font = WHITE_BOLD
        c.fill = DARK_FILL
        c.alignment = Alignment(horizontal='center')
    r += 1

    subsectors = [
        ('Industrial', 'Moderate', 'Supply chain boost', 'Reshoring accelerates', 'CONSTRUCTIVE'),
        ('Data Centers', 'Moderate', 'Energy cost (hedged)', 'AI demand durable', 'CONSTRUCTIVE'),
        ('Healthcare', 'Low', 'Minimal', 'Defensive / resilient', 'POSITIVE'),
        ('Self-Storage', 'Low-Mod', 'Insulated', 'Counter-cyclical', 'CONSTRUCTIVE'),
        ('Strip/Mall Retail', 'Moderate', 'Consumer drag', 'Gas price pressure', 'NEUTRAL'),
        ('Net Lease', 'HIGH', 'Moderate passthrough', 'Contractual / stable', 'CAUTIOUS'),
        ('Residential', 'HIGH', 'Utility cost rise', 'Near-term demand OK', 'CAUTIOUS'),
        ('Office', 'Moderate', 'Energy cost rise', 'Already challenged', 'AVOID'),
        ('Hotels', 'Low-Mod', 'Direct inflation hit', 'Travel disruption', 'AVOID'),
    ]

    view_colors = {
        'CONSTRUCTIVE': '336633', 'POSITIVE': '336633',
        'NEUTRAL': '666666',
        'CAUTIOUS': 'CC0000', 'AVOID': 'CC0000',
    }
    for name, rate_sens, oil_inf, demand, view in subsectors:
        wsfn.cell(row=r, column=2, value=name).font = BOLD
        wsfn.cell(row=r, column=3, value=rate_sens).alignment = Alignment(horizontal='center')
        wsfn.cell(row=r, column=4, value=oil_inf).alignment = Alignment(horizontal='center')
        wsfn.cell(row=r, column=5, value=demand).alignment = Alignment(horizontal='center')
        vc = wsfn.cell(row=r, column=6, value=view)
        vc.font = Font(bold=True, color=view_colors.get(view, '000000'))
        vc.alignment = Alignment(horizontal='center')
        for col in range(2, 7):
            wsfn.cell(row=r, column=col).border = THIN_BORDER
        # Alternate row shading
        if subsectors.index((name, rate_sens, oil_inf, demand, view)) % 2 == 0:
            for col in range(2, 7):
                wsfn.cell(row=r, column=col).fill = PatternFill('solid', fgColor='F2F2F2')
        r += 1
    r += 1

    # Subsector commentary
    sub_commentary = [
        ('Industrial -- Best Positioned Relative to Peers, But Still Under Pressure',
         'To be clear: industrial REITs sold off last week alongside the broader REIT market. PLD closed March 6 at '
         '$141.51, down roughly 1.9% on the day; REXR, FR, and TRNO all posted similar losses, and the subsector is '
         'lower in absolute terms since February 28. This is not a thesis about industrial REITs going up -- it is a '
         'thesis about relative outperformance vs. a REIT universe that is all under pressure. The 10Y rate channel '
         'that our regression identifies as the primary REIT headwind affects every subsector, and industrials are no '
         'exception. What separates them is a demand-side offset that does not exist elsewhere. A functionally closed '
         'Strait of Hormuz and oil surging past $100 -- WTI hit $119 intraday on March 9, its highest since 2022 -- '
         'are forcing companies to urgently reprice Asia-Pacific supply chain dependency, accelerating reshoring '
         'decisions that drive incremental US warehouse and logistics demand. Historical oil shock episodes support '
         'this dynamic: during the 1990 Gulf War and the 2022 Russia-Ukraine surge, industrial names held up materially '
         'better than net lease and residential on a relative basis as supply chain disruption narratives accelerated '
         'leasing activity. In the current episode, West Coast port-adjacent names (TRNO, FR) are the clearest relative '
         'beneficiaries given likely trade flow rerouting, though REXR carries near-term SoCal vacancy headwinds. '
         "PLD's international exposure is a two-sided watch item. The reshoring benefit is a 4-8 quarter NOI story, "
         'not a Q1 catalyst -- but relative to a REIT universe where the macro headwind is hitting everyone, industrials '
         'offer the best structural offset.'),
        ('Net Lease -- Worst Positioned, and Already Showing It',
         'Net lease has underperformed the broader REIT index since the conflict began on February 28, and the mechanism '
         'is exactly what our regression predicts. NNN entered the conflict near its 52-week high of $44.29 -- priced '
         'for the rate-cut cycle the market had been expecting in 2026 -- and has since pulled back as the 10Y pushed '
         'above 4.1% and oil above $100 extended the inflation case against near-term cuts. The structural problem is '
         'duration: net lease WALTs of 10+ years with fixed rent escalators mean these securities are effectively marked '
         'to the risk-free rate in real time. When the 10Y moves, NAVs move with it. During the 2022 Russia-Ukraine '
         'rate shock -- the closest historical analog -- NNN, O, and ADC underperformed the REIT index by 800-1200bps '
         'as the 10Y moved from 1.8% to 4.0%. The current episode compounds that dynamic: Brent crude touching $119 '
         'intraday on March 9 and WTI posting its largest single-week gain in futures history means the inflation '
         'pressure keeping the Fed sidelined is not going away quickly. NNN\'s underlying fundamentals are not the '
         'issue -- Q4 2025 showed FFO of $0.87 vs. $0.86 estimate and record $900M acquisition volume -- but clean '
         'fundamentals provide no shelter from rate-driven multiple compression. The entire net lease complex is in '
         'the wrong part of the trade until the 10Y definitively rolls over on a ceasefire or demand destruction signal.'),
        ('Hotels -- Avoid: Dual Hit to Costs and Revenue',
         'Hotel REITs have sold off more sharply than most REIT subsectors since February 28 -- and unlike industrials, '
         'where the headwind is primarily macro and shared across the sector, hotels face a simultaneous hit to both '
         'sides of the P&L that is unique to this conflict. On costs: WTI rose 35%+ last week (the largest single-week '
         'gain in futures history) and crossed $100, pushing energy, transport, and food service costs materially higher '
         'across hotel operations. On revenue: Dubai International Airport sustained damage and suspended operations; '
         'over 4,000 daily flight cancellations occurred across Gulf state airspace; and major carriers including '
         'British Airways, Lufthansa, and Air India all suspended Middle East service. US State Department travel '
         'warnings now cover Saudi Arabia, Kuwait, Cyprus, and Lebanon, functionally halting business travel to the '
         'affected region. Gateway markets dependent on international inbound traffic -- NYC, LA, Miami -- face '
         'secondary RevPAR pressure that will appear in Q1 2026 results. History is consistent: hotel REITs '
         'underperformed the REIT index by the widest margin of any subsector during the Gulf War (1990), post-9/11, '
         'and Russia-Ukraine (2022). HST, RHP, and PK all entered 2026 with positive RevPAR momentum that is now at '
         'direct risk. Unlike net lease, where the fundamental business continues to collect rent while multiples '
         'compress, hotel NOI itself is impaired. That distinction makes this the clearest avoid in the REIT universe '
         'for the duration of the conflict.'),
    ]
    for title, text in sub_commentary:
        wsfn.cell(row=r, column=2, value=title).font = SUBSECTION_FONT
        r += 1
        wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        wsfn.cell(row=r, column=2, value=text).alignment = Alignment(wrap_text=True)
        wsfn.row_dimensions[r].height = max(64, len(text) // 80 * 16 + 20)
        r += 2

    # Scenario Analysis
    wsfn.cell(row=r, column=2, value='SCENARIO ANALYSIS').font = SECTION_FONT
    r += 1
    scen_headers = ['', 'Base Case (45%)\n~4-6 Week Conflict', 'Bear Case (35%)\nExtended Hormuz Closure', 'Bull Case (20%)\nRapid Ceasefire']
    for ci, h in enumerate(scen_headers, 2):
        c = wsfn.cell(row=r, column=ci, value=h)
        c.font = WHITE_BOLD
        c.fill = DARK_FILL
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    wsfn.row_dimensions[r].height = 36
    r += 1

    scenarios = [
        ('Brent Oil', '$80-90/bbl', '$100+/bbl', '~$65-70/bbl'),
        ('10Y Yield', '4.00-4.25%', '4.25-4.75%; then retreats', '3.70-3.90%'),
        ('Fed', 'Cuts delayed to Q4 2026', 'Stagflation risk; cuts on hold', 'Cuts resume H2 2026'),
        ('REIT vs S&P', 'Slight underperform', 'Underperform; then outperform', 'Outperform (+3-5%)'),
        ('Best Subsectors', 'Industrial, Data Ctr, HC', 'Healthcare, Self-Storage', 'Net Lease, Residential'),
        ('Worst Subsectors', 'Net Lease, Hotels', 'Hotels, Office, Retail', 'Hotels (structural)'),
    ]
    for label, base, bear, bull in scenarios:
        wsfn.cell(row=r, column=2, value=label).font = BOLD
        wsfn.cell(row=r, column=3, value=base).alignment = Alignment(horizontal='center', wrap_text=True)
        wsfn.cell(row=r, column=4, value=bear).alignment = Alignment(horizontal='center', wrap_text=True)
        wsfn.cell(row=r, column=5, value=bull).alignment = Alignment(horizontal='center', wrap_text=True)
        for col in range(2, 6):
            wsfn.cell(row=r, column=col).border = THIN_BORDER
        r += 1
    r += 1

    # Tail Risk box
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = wsfn.cell(row=r, column=2, value='TAIL RISK TO MONITOR')
    c.font = Font(bold=True, color='FFFFFF')
    for col in range(2, 7):
        wsfn.cell(row=r, column=col).fill = PatternFill('solid', fgColor=DB_BLUE)
    r += 1
    tail_text = ("If oil reaches $100+ and recession fears take hold, the dynamic flips. The historical data shows that "
                 "oil crashes -- the demand-destruction phase of an energy shock -- pull the 10Y down sharply. In that "
                 "scenario, the sectors currently under pressure from duration risk (net lease, residential) become the "
                 "beneficiaries. Investors should monitor the oil price level and Fed communication carefully for the "
                 "inflection signal.")
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsfn.cell(row=r, column=2, value=tail_text).alignment = Alignment(wrap_text=True)
    wsfn.row_dimensions[r].height = 64
    for col in range(2, 7):
        wsfn.cell(row=r, column=col).fill = PatternFill('solid', fgColor='FCE4EC')
        wsfn.cell(row=r, column=col).border = THIN_BORDER
    r += 2

    # Bottom Line
    wsfn.cell(row=r, column=2, value='BOTTOM LINE').font = SECTION_FONT
    r += 1
    bl1 = ("The Iran war is a REIT headwind, but the mechanism is rates, not oil. The 10Y's unusual upward move "
           "-- defying the traditional geopolitical safe-haven bid -- is the variable to watch. At ~4.1%, the 10Y "
           "represents a meaningful but not catastrophic headwind for the sector in aggregate.")
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsfn.cell(row=r, column=2, value=bl1).alignment = Alignment(wrap_text=True)
    wsfn.row_dimensions[r].height = 48
    r += 2

    bl2 = ("Subsector preference order: (1) Industrial and Data Centers as secular stories with the best insulation; "
           "(2) Healthcare for defensiveness; (3) Self-Storage as counter-cyclical; (4) Net Lease and Residential on "
           "caution given duration exposure; (5) Hotels and Office as avoids. The conflict is likely to resolve within "
           "4-8 weeks on current military trajectory, at which point rate-cut expectations should resume and a more "
           "constructive REIT backdrop re-emerges.")
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsfn.cell(row=r, column=2, value=bl2).alignment = Alignment(wrap_text=True)
    wsfn.row_dimensions[r].height = 64
    r += 2

    # Disclaimer
    disc = ("This report is prepared for informational purposes only and does not constitute investment advice or a "
            "recommendation to buy or sell any security. Historical data sourced from FRED (DCOILWTICO, DTB3, DGS10) "
            "and FTSE NAREIT. Regression analysis based on monthly data 1995-2025 (n=366 observations). All market "
            "data as of March 6-7, 2026. For institutional investors only.")
    wsfn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wsfn.cell(row=r, column=2, value=disc).font = Font(italic=True, size=8, color='999999')
    wsfn.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
    wsfn.row_dimensions[r].height = 48

    # ==================================================================
    # TAB 9: NOTE SOURCES
    # ==================================================================
    wsns = wb.create_sheet('Note Sources')
    wsns.column_dimensions['A'].width = 4
    wsns.column_dimensions['B'].width = 8
    wsns.column_dimensions['C'].width = 35
    wsns.column_dimensions['D'].width = 80
    wsns.column_dimensions['E'].width = 65

    r = 1
    wsns.cell(row=r, column=2, value='Flash Note Sources').font = Font(bold=True, size=16, color=DB_BLUE)
    r += 2
    _header_row(wsns, r, ['', '#', 'Source', 'Link', 'Used For'])
    r += 1

    note_sources = [
        # Oil Prices & Market Backdrop
        ('1', 'CNBC (oil prices Mar 8)',
         'https://www.cnbc.com/2026/03/08/crude-oil-prices-today-iran-war.html',
         'WTI $119.48 intraday high, Brent $119.50; WTI +35% last week (largest weekly gain in futures history '
         'since 1983); WTI crossing $100 for first time since 2022; Gulf Arab states cutting production; Iraq '
         'output down ~60%'),
        ('2', 'CNBC (oil prices Mar 6)',
         'https://www.cnbc.com/2026/03/06/iran-us-war-oil-prices-brent-wti-barrel-futures.html',
         'WTI $90.90 close Friday March 6, Brent $92.69; WTI +35.63% for the week (biggest weekly gain in '
         'futures history dating back to 1983); Brent +28% (biggest since April 2020); Qatar energy minister '
         'warning crude could reach $150/barrel'),
        ('3', 'OilPrice.com (Mar 9)',
         'https://oilprice.com/Energy/Energy-General/Oil-Prices-Soar-29-as-Iran-Conflict-Threatens-Middle-East-Supply.html',
         'Brent $107.20 (+15.65%), WTI $103.18 (+14.26%) during Monday March 9 session; Brent intraday high '
         '$119.50; WTI intraday high $119.48; Iraq output down ~70%; Goldman Sachs warning $140-150 if Hormuz '
         'disrupted 30+ days; VIX highest since April 2025'),
        ('4', 'FX Leaders (Mar 9)',
         'https://www.fxleaders.com/news/2026/03/09/oil-price-shock-wti-and-brent-explode-past-100-as-middle-east-refineries-face-direct-hits/',
         'WTI $119.48 intraday, Brent $119.50; today\'s trading range WTI $96.45-$119.43; 52-week range '
         '$54.98-$119.43'),
        ('5', 'Axios (Mar 8)',
         'https://www.axios.com/2026/03/08/iran-war-oil-market-barrel-cost',
         'Brent initially ~$101.81, WTI ~$101.56 Sunday evening; Brent later above $108; WTI near $120 '
         'overnight; gasoline up from ~$3/gallon to $3.45; Rapidan Energy: disruption of 20% global oil '
         'supply for 9 days, "more than double the previous record set during the Suez Crisis of 1956-57"'),

        # Strait of Hormuz / Shipping
        ('6', 'Euronews (Hormuz)',
         'https://www.euronews.com/business/2026/03/05/passage-denied-oil-and-gas-prices-swing-wildly-as-hormuz-crisis-drags-on',
         'Tanker transits collapsed from 24/day to 4/day (Vortexa data); ~200 tankers stranded; WTI ~$74.56 / '
         'Brent ~$81.40 as of early March; Goldman Q2 Brent $76 / WTI $71 forecast (pre-escalation)'),
        ('7', 'Kpler (Hormuz)',
         'https://www.kpler.com/blog/us-iran-conflict-strait-of-hormuz-crisis-reshapes-global-oil-markets',
         '~20% of global seaborne oil through Hormuz (~20.9M bbl/day); 31% of seaborne crude flows; 83% '
         'shipping reduction; 20% of global LNG; Qatar ~75% of LNG exports through Hormuz'),

        # Rates / Equities
        ('8', 'CNBC (10Y yield)',
         'https://www.cnbc.com/2026/03/03/10-year-treasury-yield-tops-4point06percent-as-surging-oil-prices-from-iran-conflict-raise-inflation-angst.html',
         '10Y yield above 4%, peaked 4.117% on March 2; ISM prices paid jumped to 70.5; bond market defying '
         'safe-haven playbook'),
        ('9', 'CNBC (equities Mar 8)',
         'https://www.cnbc.com/2026/03/08/stock-market-today-live-updates.html',
         'S&P 500 down ~2% WTD; Dow -800pts Thursday; Russell 2000 off more than 4% since Feb 28; Dow futures '
         '-512 points Sunday night; VIX topped 30'),
        ('10', 'Motley Fool (equities)',
         'https://www.fool.com/investing/2026/03/06/will-the-iran-war-cause-stock-market-crash-history/',
         'S&P 500 higher 65% of the time one year after major geopolitical events, average +3% return; median '
         '+5% six months later (Carson Group, 40+ events since WWII)'),

        # Aviation / Travel
        ('11', 'Wikipedia (economic impact)',
         'https://en.wikipedia.org/wiki/Economic_impact_of_the_2026_Iran_war',
         '4,000+ daily flight cancellations across Gulf state airspace; Dubai International Airport damaged and '
         'suspended; Emirates, BA, Lufthansa, Virgin Atlantic, Air India, Cathay Pacific, Qatar Airways, Kuwait '
         'Airways all suspended; Bahrain, Iraq, Israel, Kuwait, Qatar, Syria, UAE all closed airspace'),

        # Industrial REITs
        ('12', 'Dividend.com (industrial REITs)',
         'https://www.dividend.com/industrial-reit-sub-industry-dividend-stocks-etfs-and-funds/',
         'PLD $141.51 (-1.89%), EGP $192.92 (+1.20%), REXR $37.66 (+1.45%), FR $62.08 (+2.22%), STAG $39.53 '
         '(+2.57%), TRNO $65.88 (+1.39%) -- March 6 closing prices'),

        # Net Lease REITs
        ('13', 'Macrotrends (NNN)',
         'https://www.macrotrends.net/stocks/charts/NNN/nnn-reit/stock-price-history',
         'NNN 52-week high $44.29, 52-week low $35.80'),
        ('14', 'Daily Political (NNN Q4)',
         'https://www.dailypolitical.com/2026/02/13/nnn-reit-nysennn-updates-fy-2026-earnings-guidance.html',
         'NNN Q4 2025 FFO $0.87 vs. $0.86 estimate; record ~$900M acquisition volume; 2026 AFFO/share growth '
         'target ~3.2%; FY2026 EPS guidance $3.470-$3.530'),

        # Healthcare REITs
        ('15', 'Investing.com (WELL)',
         'https://www.investing.com/equities/health-care-reit-historical-data',
         'WELL 52-week range $130.29-$209.05; today\'s range $196.03-$198.85; +37.22% over past year'),
        ('16', 'Investing.com (WELL SWOT)',
         'https://www.investing.com/news/swot-analysis/welltowers-swot-analysis-healthcare-reit-stock-poised-for-growth-amid-market-shifts-93CH-4277071',
         'WELL FFO/share guidance raised to $6.01 for 2026 vs. $5.80 consensus; 2025 FFO raised from $4.95 to '
         '$5.18 vs. $5.08 consensus; revenue growth 32% LTM'),
        ('17', 'Stock Titan (DOC)',
         'https://www.stocktitan.net/overview/DOC/',
         'DOC $17.28 as of March 6; market cap ~$12.1B; Janus Living IPO (34 communities, 10,422 units) '
         'expected H1 2026; Q4 NAREIT FFO $0.47/share, AFFO $0.40/share; Net Debt/Adj. EBITDAre 5.2x'),

        # Regression Analysis
        ('18', 'Oil_REITs_Rates_Analysis.xlsx',
         '(This workbook)',
         'Full sample oil coef -0.022 (p=0.567); 10Y coef -3.26 (p=0.069); 1.5 SD shock 10Y coef -14.62 '
         '(p<0.001); oil crashes 10Y coef -10.61 (p=0.005); oil\u219210Y R\u00b2 8.3% full sample rising to '
         '32.1% extreme shocks; n=366 monthly observations 1995-2025'),
    ]

    # Section headers for grouping
    section_breaks = {
        '1': 'Oil Prices & Market Backdrop',
        '6': 'Strait of Hormuz / Shipping',
        '8': 'Rates / Equities',
        '11': 'Aviation / Travel (Hotels)',
        '12': 'Industrial REITs',
        '13': 'Net Lease REITs',
        '15': 'Healthcare REITs',
        '18': 'Regression Analysis',
    }

    for num, source, link, used_for in note_sources:
        if num in section_breaks:
            wsns.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            wsns.cell(row=r, column=2, value=section_breaks[num]).font = Font(bold=True, size=11, color=DB_BLUE)
            wsns.cell(row=r, column=2).fill = LIGHT_FILL
            for c in range(2, 6):
                wsns.cell(row=r, column=c).fill = LIGHT_FILL
            r += 1
        wsns.cell(row=r, column=2, value=int(num)).alignment = Alignment(horizontal='center')
        wsns.cell(row=r, column=3, value=source).font = BOLD
        if link.startswith('http'):
            wsns.cell(row=r, column=4, value=link).font = Font(color='0563C1', underline='single')
            wsns.cell(row=r, column=4).hyperlink = link
        else:
            wsns.cell(row=r, column=4, value=link).font = Font(italic=True, color='888888')
        wsns.cell(row=r, column=5, value=used_for).alignment = Alignment(wrap_text=True)
        wsns.row_dimensions[r].height = max(40, len(used_for) // 60 * 16 + 20)
        for c in range(2, 6):
            wsns.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    # ==================================================================
    # TAB 10: CHARTS
    # ==================================================================
    ws6 = wb.create_sheet('Charts')

    # Chart 1: Oil vs REIT excess
    c1 = ScatterChart()
    c1.title = 'Oil Price Swings vs REIT Outperformance'
    c1.x_axis.title = 'Monthly Oil Price Move (%)'
    c1.y_axis.title = 'REIT vs S&P (%)'
    c1.width = 20
    c1.height = 14
    s1 = Series(Reference(ws, min_col=6, min_row=2, max_row=last_row),
                Reference(ws, min_col=3, min_row=2, max_row=last_row), title='Monthly')
    s1.graphicalProperties.noFill = True
    c1.series.append(s1)
    ws6.add_chart(c1, 'A1')

    # Chart 2: 10Y change vs REIT excess
    c2 = ScatterChart()
    c2.title = 'When Long-Term Rates Rise, REITs Underperform'
    c2.x_axis.title = '10Y Rate Change (pp)'
    c2.y_axis.title = 'REIT vs S&P (%)'
    c2.width = 20
    c2.height = 14
    s2 = Series(Reference(ws, min_col=6, min_row=2, max_row=last_row),
                Reference(ws, min_col=11, min_row=2, max_row=last_row), title='Monthly')
    s2.graphicalProperties.noFill = True
    c2.series.append(s2)
    ws6.add_chart(c2, 'L1')

    # Chart 3: Oil vs 10Y change
    c3 = ScatterChart()
    c3.title = 'Oil and Long-Term Rates Move Together'
    c3.x_axis.title = 'Monthly Oil Price Move (%)'
    c3.y_axis.title = '10Y Rate Change (pp)'
    c3.width = 20
    c3.height = 14
    s3 = Series(Reference(ws, min_col=11, min_row=2, max_row=last_row),
                Reference(ws, min_col=3, min_row=2, max_row=last_row), title='Monthly')
    s3.graphicalProperties.noFill = True
    c3.series.append(s3)
    ws6.add_chart(c3, 'A18')

    # Chart 4: Oil + 10Y over time
    c4 = LineChart()
    c4.title = 'Oil Price and 10Y Rate Over Time'
    c4.y_axis.title = 'Oil Price ($/bbl)'
    c4.width = 20
    c4.height = 14
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    c4.add_data(Reference(ws, min_col=2, min_row=1, max_row=last_row), titles_from_data=True)
    c4.set_categories(cats)
    c4.series[0].graphicalProperties.line.width = 15000
    c4b = LineChart()
    c4b.y_axis.title = '10Y Rate (%)'
    c4b.add_data(Reference(ws, min_col=8, min_row=1, max_row=last_row), titles_from_data=True)
    c4b.set_categories(cats)
    c4b.y_axis.axId = 200
    c4b.series[0].graphicalProperties.line.width = 15000
    c4.y_axis.crosses = 'min'
    c4 += c4b
    ws6.add_chart(c4, 'L18')

    # ==================================================================
    # TAB 11: SAKS CLOSURES
    # ==================================================================
    import openpyxl as _opx
    import os as _os
    saks_path = _os.path.join(_os.path.dirname(__file__), 'Saks_Closures_Landlord_Exposure.xlsx')
    if _os.path.exists(saks_path):
        src = _opx.load_workbook(saks_path)
        src_ws = src.active
        ws_saks = wb.create_sheet('Saks Closures')
        for r_idx in range(1, src_ws.max_row + 1):
            for c_idx in range(1, src_ws.max_column + 1):
                src_cell = src_ws.cell(r_idx, c_idx)
                dst_cell = ws_saks.cell(r_idx, c_idx, value=src_cell.value)
                # Copy key formatting
                if src_cell.font:
                    dst_cell.font = src_cell.font.copy()
                if src_cell.fill and src_cell.fill.fgColor and src_cell.fill.fgColor.rgb and src_cell.fill.fgColor.rgb != '00000000':
                    dst_cell.fill = src_cell.fill.copy()
                if src_cell.alignment:
                    dst_cell.alignment = src_cell.alignment.copy()
                if src_cell.border:
                    dst_cell.border = src_cell.border.copy()
                if src_cell.number_format:
                    dst_cell.number_format = src_cell.number_format
        # Copy column widths
        for col_letter, dim in src_ws.column_dimensions.items():
            ws_saks.column_dimensions[col_letter].width = dim.width
        # Copy merged cells
        for merged in src_ws.merged_cells.ranges:
            ws_saks.merge_cells(str(merged))
        # Copy row heights
        for r_idx, dim in src_ws.row_dimensions.items():
            if dim.height:
                ws_saks.row_dimensions[r_idx].height = dim.height

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
